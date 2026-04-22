import os
import json
import base64
import datetime
import logging
import threading
import socket
import re
import shutil
import tempfile
import queue
import time
import cv2
import pandas as pd
import openpyxl
from pathlib import Path
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from django.views.decorators.csrf import csrf_exempt
from google import genai
from google.genai import types
from django.core.cache import cache
from rapidfuzz import process, fuzz
from htr.engine import HTREngine

# Configure Logging
logger = logging.getLogger(__name__)

# --- EXCEL WRITE QUEUE SYSTEM (Hội đồng Tứ tướng's Ultimate Fix) ---
excel_write_queue = queue.Queue()
excel_lock = threading.Lock()

def excel_worker():
    """Worker duy nhất chuyên trách việc ghi file Excel để tránh tranh chấp."""
    logger.info("🚀 [EXCEL-WORKER] Started and waiting for tasks...")
    while True:
        try:
            task = excel_write_queue.get()
            if task is None: break
            
            func, args, kwargs = task
            logger.info(f"📥 [EXCEL-WORKER] Processing task: {func.__name__}")
            func(*args, **kwargs)
            logger.info(f"✅ [EXCEL-WORKER] Finished task: {func.__name__}")
        except Exception as e:
            logger.error(f"💥 [EXCEL-WORKER] Fatal error in loop: {str(e)}")
        finally:
            excel_write_queue.task_done()

# Khởi động Worker ngay khi nạp module
threading.Thread(target=excel_worker, daemon=True).start()

# AI Processing Pool to prevent Rate Limit and Server Crash
ai_executor = ThreadPoolExecutor(max_workers=5)

def get_session_config():
    return cache.get("global_session_config", {
        "excel_filename": None,
        "subject": None,
        "roster_json": None,
        "sheet_name": None,
        "mapping": {}
    })

def set_session_config(config):
    cache.set("global_session_config", config, timeout=86400)

def add_to_scan_store(session_id, image_b64):
    key = f"scan_store_{session_id}"
    imgs = cache.get(key, [])
    imgs.append(image_b64)
    cache.set(key, imgs, timeout=86400)

def pop_scan_store(session_id):
    key = f"scan_store_{session_id}"
    imgs = cache.get(key, [])
    if imgs:
        cache.delete(key)
    return imgs

# Load environment variables
load_dotenv()

# Kiểm tra Gemini API Key
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    logger.warning("⚠️ GEMINI_API_KEY not found in environment variables!")

def get_backend_root():
    return Path(__file__).resolve().parent.parent

def normalize_string(s):
    """Chuẩn hóa chuỗi tiếng Việt: viết thường, bỏ dấu, xóa khoảng trắng thừa."""
    if not s: return ""
    s = str(s).lower().strip()
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[đ]', 'd', s)
    s = re.sub(r'\s+', ' ', s)
    return s

def parse_score(raw_score):
    """🎯 Chuyển đổi điểm số từ AI sang dạng số thực (float) an toàn."""
    if raw_score is None or str(raw_score).strip() == "":
        return None
    try:
        # Xử lý trường hợp AI trả về chuỗi có dấu phẩy hoặc ký tự lạ
        s = str(raw_score).replace(',', '.').strip()
        # Chỉ giữ lại số và dấu chấm
        s = re.sub(r'[^\d.]', '', s)
        if not s: return None
        val = float(s)
        # Giới hạn điểm từ 0-10 (nếu cần)
        return max(0.0, min(10.0, val))
    except Exception:
        return None

def find_best_student_match(gemini_id, gemini_name, roster_data):
    """
    🎯 Sử dụng RapidFuzz để tìm học sinh khớp nhất từ danh sách Roster.
    roster_data: list of dicts [{"id": "...", "name": "..."}]
    """
    if not roster_data:
        return gemini_id, gemini_name, False, 0

    # 🛡️ Issue 3: Handle null response
    gemini_id = gemini_id or ""
    gemini_name = gemini_name or ""

    best_match = None
    highest_score = 0

    norm_gemini_id = str(gemini_id).strip().lower()
    norm_gemini_name = normalize_string(gemini_name)

    # 1. Thử khớp theo ID trước (Ưu tiên tuyệt đối)
    if norm_gemini_id:
        for student in roster_data:
            s_id = str(student.get('id', '')).strip().lower()
            if not s_id: continue
            
            # Khớp chính xác hoặc score rất cao
            score = fuzz.ratio(norm_gemini_id, s_id)
            if score > 90:
                return student.get('id'), student.get('name'), True, score

    # 2. Thử khớp theo Tên nếu ID không khớp tốt
    if norm_gemini_name:
        for student in roster_data:
            s_name = normalize_string(student.get('name', ''))
            if not s_name: continue
            
            score = fuzz.ratio(norm_gemini_name, s_name)
            if score > highest_score:
                highest_score = score
                best_match = student

    # Ngưỡng chấp nhận cho khớp tên là 80%
    if highest_score >= 80 and best_match:
        return best_match.get('id'), best_match.get('name'), True, highest_score

    # Không khớp được thì trả về kết quả thô ban đầu
    return gemini_id, gemini_name, False, highest_score

def secure_filename(filename):
    """Dọn sạch ký tự nguy hiểm (Path Traversal) khỏi tên file."""
    import os
    if not filename: return ""
    # Chỉ lấy phần tên, bỏ qua mọi đường dẫn
    clean_name = os.path.basename(str(filename))
    # Loại bỏ triệt để thư mục cha và dấu nháy
    clean_name = clean_name.replace("..", "").replace("/", "").replace("\\", "")
    return clean_name

def extract_base64_data(image_data_base64):
    try:
        if not image_data_base64: return None
        imgstr = image_data_base64.split(';base64,')[1] if ';' in image_data_base64 else image_data_base64
        return base64.b64decode(imgstr)
    except Exception as e:
        logger.error(f"Error decoding base64: {str(e)}")
        return None

def _internal_save_score(excel_filename, sheet_name, student_id, score, id_col_idx=None, score_col_idx=None, data_row_start=2):
    """Lõi xử lý ghi Excel: Chạy trong Worker Thread, đảm bảo tuần tự.
    Hỗ trợ 2 chế độ:
    - Nếu có id_col_idx + score_col_idx: dùng trực tiếp (từ frontend mapping)
    - Nếu không: fallback sang session mapping (tên cột)
    """
    with excel_lock:
        temp_path = None
        try:
            file_path = get_backend_root() / 'media' / 'rosters' / excel_filename
            if not file_path.exists():
                logger.warning(f"⚠️ [SAVE-SCORE] File not found: {excel_filename}")
                return

            fd, temp_path = tempfile.mkstemp(suffix='.xlsx', dir=str(file_path.parent))
            os.close(fd)

            wb = openpyxl.load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                logger.warning(f"⚠️ [SAVE-SCORE] Sheet '{sheet_name}' not found in {excel_filename}. Available: {wb.sheetnames}")
                wb.close()
                if os.path.exists(temp_path): os.remove(temp_path)
                return

            sheet = wb[sheet_name]

            # Xác định cột ID và Score
            final_id_col = id_col_idx
            final_score_col = score_col_idx

            if not final_id_col or not final_score_col:
                # Fallback: dùng session mapping (tên cột)
                config = get_session_config()
                mapping = config.get("mapping", {})
                
                # Check if mapping has column indices (from frontend)
                if mapping.get("idCol") and mapping.get("scoreCol"):
                    final_id_col = int(mapping["idCol"])
                    final_score_col = int(mapping["scoreCol"])
                    data_row_start = int(mapping.get("dataRowStart", 2))
                    logger.info(f"📊 [SAVE-SCORE] Using index mapping: id_col={final_id_col}, score_col={final_score_col}, data_start={data_row_start}")
                elif mapping.get("id") and mapping.get("score"):
                    # Legacy: tên cột
                    headers = [cell.value for cell in sheet[1]]
                    try:
                        final_id_col = headers.index(mapping["id"]) + 1
                        final_score_col = headers.index(mapping["score"]) + 1
                    except ValueError:
                        logger.warning(f"⚠️ [SAVE-SCORE] Column names not found in headers. mapping={mapping}, headers={headers}")
                        wb.close()
                        if os.path.exists(temp_path): os.remove(temp_path)
                        return
                else:
                    logger.warning(f"⚠️ [SAVE-SCORE] No valid mapping found! mapping={mapping}. Ghi điểm bị bỏ qua.")
                    wb.close()
                    if os.path.exists(temp_path): os.remove(temp_path)
                    return

            # Parse score to float
            valid_score = parse_score(score)
            if valid_score is None:
                logger.warning(f"⚠️ [SAVE-SCORE] Invalid score skipped: '{score}' for student {student_id}")
                wb.close()
                if os.path.exists(temp_path): os.remove(temp_path)
                return

            # Tìm và ghi điểm
            found = False
            for row in range(data_row_start, sheet.max_row + 1):
                cell_id = sheet.cell(row=row, column=final_id_col).value
                if str(cell_id).strip() == str(student_id).strip():
                    sheet.cell(row=row, column=final_score_col).value = valid_score
                    found = True
                    break

            if found:
                wb.save(temp_path)
                wb.close()
                os.replace(temp_path, str(file_path))
                logger.info(f"🎯 [SAVE-SCORE] ✅ Updated student '{student_id}' → score={valid_score} in '{sheet_name}'")
            else:
                logger.warning(f"⚠️ [SAVE-SCORE] Student '{student_id}' not found in sheet '{sheet_name}' (searched rows {data_row_start}-{sheet.max_row}, id_col={final_id_col})")
                wb.close()
                if os.path.exists(temp_path): os.remove(temp_path)
        except Exception as e:
            logger.error(f"💥 [SAVE-SCORE] Fatal error: {str(e)}", exc_info=True)
            if temp_path and os.path.exists(temp_path): os.remove(temp_path)

def update_excel_score(excel_filename, sheet_name, student_id, score, id_col_idx=None, score_col_idx=None, data_row_start=2):
    """Đẩy yêu cầu ghi điểm vào hàng đợi và phản hồi ngay lập tức."""
    if not excel_filename or not student_id:
        logger.warning(f"⚠️ [UPDATE-SCORE] Skipped: excel_filename={excel_filename}, student_id={student_id}")
        return False
    excel_write_queue.put((_internal_save_score, (excel_filename, sheet_name, student_id, score, id_col_idx, score_col_idx, data_row_start), {}))
    return True

def retry_ai_call(max_retries=3, delay=2):
    """Decorator thực hiện thử lại cuộc gọi AI nếu thất bại."""
    def decorator(func):
        def wrapper(*args, **kwargs):
            for i in range(max_retries):
                try:
                    res = func(*args, **kwargs)
                    if res and res.get('status') == 'success':
                        return res
                    logger.warning(f"⚠️ [RETRY] {func.__name__} attempt {i+1} returned invalid data.")
                except Exception as e:
                    logger.warning(f"⚠️ [RETRY] {func.__name__} attempt {i+1} failed: {str(e)}")
                
                if i < max_retries - 1:
                    time.sleep(delay * (2 ** i)) # Exponential backoff
            return None
        return wrapper
    return decorator

@retry_ai_call(max_retries=3)
def call_gemini_native(image_filename, excel_filename, subject, roster_json):
    """🎯 GEMINI API: Sử dụng Google GenAI SDK với Prompt tối ưu (Issue 6)."""
    if not GEMINI_API_KEY:
        return {"status": "error", "msg": "API Key missing"}

    try:
        image_path = (get_backend_root() / 'media' / 'scanned_images' / image_filename).absolute()
        if not image_path.exists(): return None

        with open(image_path, "rb") as f: image_bytes = f.read()

        prompt = (
            f"Nhiệm vụ: Chấm điểm bài thi môn {subject}. "
            f"Phân tích ảnh và trích xuất: 1. Tên/ID học sinh, 2. Điểm số. "
            f"Danh sách tham khảo: {roster_json}. "
            f"CHỈ TRẢ VỀ JSON DUY NHẤT, KHÔNG CÓ TEXT THỪA, KHÔNG MARKDOWN. "
            f"Format: {{\"studentId\": \"...\", \"studentName\": \"...\", \"score\": \"...\", \"status\": \"success\"}}"
        )

        client = genai.Client(api_key=GEMINI_API_KEY)
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[prompt, types.Part.from_bytes(data=image_bytes, mime_type="image/jpeg")],
            config=types.GenerateContentConfig(response_mime_type="application/json", temperature=0.0)
        )

        return json.loads(response.text.strip())
    except Exception as e:
        logger.error(f"⚠️ [GEMINI-API] Error: {str(e)}")
        raise e

# --- API ENDPOINTS ---

@api_view(['GET'])
@permission_classes([AllowAny])
def health_check(request):
    return Response({
        "status": "healthy",
        "mode": "Gemini API SDK + Pandas Core + OpenPyXL",
        "api_key_configured": bool(GEMINI_API_KEY)
    })

@api_view(['POST'])
@csrf_exempt
@permission_classes([AllowAny])
def process_test_paper(request):
    img_list = request.data.get('image_data_list', [])
    if not img_list and request.data.get('image_data'):
        img_list = [request.data.get('image_data')]

    subject = request.data.get('expected_subject')
    excel_filename = secure_filename(request.data.get('excel_filename'))
    roster_data = request.data.get('roster', [])
    roster_str = json.dumps(roster_data, ensure_ascii=False)
    mapping_config = request.data.get('mapping_config')  # Frontend MappingConfig

    # Sync session config
    config = get_session_config()
    if excel_filename: config["excel_filename"] = excel_filename
    if subject:
        config["subject"] = subject
        config["sheet_name"] = subject  # Bug #3 fix: sync sheet_name
    if roster_str: config["roster_json"] = roster_str
    
    # Bug #3 fix: sync mapping from frontend (column indices)
    if mapping_config:
        config["mapping"] = mapping_config
        logger.info(f"📊 [PROCESS-PAPER] Synced mapping_config: {mapping_config}")
    set_session_config(config)

    # Extract column indices for direct Excel writing
    id_col_idx = mapping_config.get('idCol') if mapping_config else None
    score_col_idx = mapping_config.get('scoreCol') if mapping_config else None
    data_row_start = mapping_config.get('dataRowStart', 2) if mapping_config else 2
    sheet_name = config.get("sheet_name", "")

    logger.info(f"🚀 [PROCESS-PAPER] Processing {len(img_list)} images. excel={excel_filename}, sheet={sheet_name}, id_col={id_col_idx}, score_col={score_col_idx}")

    try:
        final_results = []
        for img_b64 in img_list:
            raw_bytes = extract_base64_data(img_b64)
            if not raw_bytes: continue

            img_name = f"scan_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
            save_path = get_backend_root() / 'media' / 'scanned_images' / img_name
            save_path.parent.mkdir(parents=True, exist_ok=True)
            with open(save_path, 'wb') as f: f.write(raw_bytes)

            res = call_gemini_native(img_name, excel_filename, subject, roster_str)
            
            if res:
                # 🎯 Thực hiện Fuzzy Matching để chuẩn hoá ID học sinh
                matched_id, matched_name, is_match, score = find_best_student_match(
                    res.get('studentId'), res.get('studentName'), roster_data
                )
                
                if is_match:
                    res['studentId'] = matched_id
                    res['studentName'] = matched_name
                    res['isFuzzyMatch'] = True
                    res['fuzzyScore'] = score
                    logger.info(f"🎯 [PROCESS-PAPER] Fuzzy matched: '{res.get('studentId')}' → '{matched_id}' (score={score})")
                else:
                    res['isFuzzyMatch'] = False
                    res['fuzzyScore'] = score
                    logger.warning(f"⚠️ [PROCESS-PAPER] No fuzzy match for ID='{res.get('studentId')}', Name='{res.get('studentName')}' (best_score={score})")

                # Tự động ghi vào Excel nếu có thông tin (dùng ID đã được map chuẩn)
                if res.get('studentId') and res.get('score'):
                    res['excelUpdated'] = update_excel_score(
                        excel_filename, sheet_name, res['studentId'], res['score'],
                        id_col_idx=id_col_idx, score_col_idx=score_col_idx, data_row_start=data_row_start
                    )
                else:
                    logger.warning(f"⚠️ [PROCESS-PAPER] Skipped Excel write: studentId={res.get('studentId')}, score={res.get('score')}")

                res['image_url'] = f"/media/scanned_images/{img_name}"
                final_results.append(res)

        return Response(final_results)
    except Exception as e:
        logger.error(f"❌ [PROCESS-PAPER] Lỗi: {str(e)}", exc_info=True)
        return Response({"error": str(e)}, status=500)

@api_view(['POST'])
@permission_classes([AllowAny])
def upload_roster_excel(request):
    file_obj = request.FILES.get('file')
    if not file_obj: return Response({"error": "No file"}, status=400)
    try:
        filename = f"roster_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(file_obj.name)}"
        save_path = get_backend_root() / 'media' / 'rosters' / filename
        save_path.parent.mkdir(parents=True, exist_ok=True)
        with open(save_path, 'wb+') as f:
            for chunk in file_obj.chunks(): f.write(chunk)

        # Update session config
        config = get_session_config()
        config["excel_filename"] = filename
        set_session_config(config)

        # Kích hoạt xử lý lại các ảnh đang chờ (nếu có trong scan_store)
        # Ở đây ta có thể quét thư mục scanned_images để xử lý những ảnh chưa có điểm

        return Response({"success": True, "filename": filename})
    except Exception as e: return Response({"error": str(e)}, status=500)

@api_view(['GET'])
@permission_classes([AllowAny])
def preview_excel(request):
    filename = secure_filename(request.query_params.get('filename'))
    sheet_name = request.query_params.get('sheet')
    if not filename: return Response({"error": "Filename required"}, status=400)

    # 🎯 Chiến thuật Shadow Copy: Copy file ra bản nháp để đọc, tránh tranh chấp với AI worker
    temp_preview_path = None
    try:
        file_path = get_backend_root() / 'media' / 'rosters' / filename
        if not file_path.exists(): return Response({"error": "File not found"}, status=404)

        # 1. Tạo file tạm riêng cho Preview
        fd, temp_preview_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        
        # 2. Copy dữ liệu sang file tạm (Dùng Lock để đảm bảo copy lúc file gốc đang rảnh)
        with excel_lock:
            shutil.copy2(str(file_path), temp_preview_path)

        # 3. Đọc dữ liệu từ file tạm bằng pandas
        xl = pd.ExcelFile(temp_preview_path)
        sheets = xl.sheet_names
        target_sheet = sheet_name if sheet_name in sheets else sheets[0]
        
        config = get_session_config()
        config["sheet_name"] = target_sheet
        set_session_config(config)

        df = pd.read_excel(temp_preview_path, sheet_name=target_sheet).fillna("")
        
        # Dọn dẹp file tạm ngay sau khi đọc xong vào RAM
        xl.close()
        if os.path.exists(temp_preview_path): os.remove(temp_preview_path)

        return Response({
            "success": True,
            "sheets": sheets,
            "currentSheet": target_sheet,
            "previewData": df.to_dict(orient='records'),
            "columns": list(df.columns)
        })
    except Exception as e:
        logger.error(f"❌ [PREVIEW-EXCEL] Error: {str(e)}")
        if temp_preview_path and os.path.exists(temp_preview_path):
            os.remove(temp_preview_path)
        return Response({"error": str(e)}, status=500)

@api_view(['POST'])
@permission_classes([AllowAny])
def analyze_excel_columns(request):
    filename = secure_filename(request.data.get('filename'))
    sheet_name = request.data.get('sheet')
    if not filename: return Response({"error": "No filename"}, status=400)

    try:
        file_path = get_backend_root() / 'media' / 'rosters' / filename
        df = pd.read_excel(file_path, sheet_name=sheet_name).fillna("")
        cols = [str(c).lower() for c in df.columns]

        mapping = {"id": "", "name": "", "score": ""}
        for i, c in enumerate(cols):
            if any(x in c for x in ["mã", "msv", "id", "stt"]): mapping["id"] = df.columns[i]
            if any(x in c for x in ["tên", "họ", "name"]): mapping["name"] = df.columns[i]
            if any(x in c for x in ["điểm", "score", "grade"]): mapping["score"] = df.columns[i]

        config = get_session_config()
        config["mapping"] = mapping # Save to session
        set_session_config(config)
        return Response({"success": True, "mapping": mapping, "columns": list(df.columns)})
    except Exception as e:
        return Response({"error": str(e)}, status=500)

@api_view(['POST'])
@permission_classes([AllowAny])
def sync_roster(request):
    """Lưu dữ liệu roster từ frontend ngược lại file Excel AN TOÀN."""
    filename = secure_filename(request.data.get('filename'))
    sheet_name = request.data.get('sheet')
    roster_data = request.data.get('roster')

    if not all([filename, sheet_name, roster_data]):
        return Response({"error": "Missing data"}, status=400)

    with excel_lock:
        temp_path = None
        try:
            file_path = get_backend_root() / 'media' / 'rosters' / filename
            if not file_path.exists(): return Response({"error": "File not found"}, status=404)

            # 1. Tạo file tạm
            fd, temp_path = tempfile.mkstemp(suffix='.xlsx', dir=str(file_path.parent))
            os.close(fd)

            # 2. Load workbook và thay thế sheet
            wb = openpyxl.load_workbook(file_path)
            
            if sheet_name in wb.sheetnames:
                std = wb[sheet_name]
                wb.remove(std)
            
            sheet = wb.create_sheet(sheet_name)
            
            # Ghi header và data
            if roster_data:
                headers = list(roster_data[0].keys())
                for c_idx, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=c_idx).value = header
                
                for r_idx, row_dict in enumerate(roster_data, 2):
                    for c_idx, header in enumerate(headers, 1):
                        sheet.cell(row=r_idx, column=c_idx).value = row_dict.get(header)

            # 3. Lưu vào file tạm
            wb.save(temp_path)
            wb.close()

            # 4. Tráo đổi (Atomic swap)
            os.replace(temp_path, str(file_path))
            
            return Response({"success": True, "msg": "Đã đồng bộ file Excel thành công"})
        except Exception as e:
            logger.error(f"❌ [SYNC-ROSTER] Error: {str(e)}")
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
            return Response({"error": str(e)}, status=500)

@api_view(['POST'])
@permission_classes([AllowAny])
def reset_system(request):
    set_session_config({"excel_filename": None, "subject": None, "roster_json": None, "sheet_name": None, "mapping": {}})
    return Response({"success": True})

# --- WebSocket Helpers & Scan Endpoints ---
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer

def send_ws_update(session_id, message_type, data):
    channel_layer = get_channel_layer()
    if channel_layer:
        async_to_sync(channel_layer.group_send)(
            f"scan_{session_id}",
            {"type": "scan_message", "message": {"type": message_type, "payload": data}}
        )

def background_ai_task(image_filename, config, session_id):
    try:
        send_ws_update(session_id, "ai_status", {"status": "analyzing", "msg": "🔍 Đang quét bài bằng Local HTR..."})
        
        # 1. 100% Local HTR Pipeline (Không dùng Gemini nữa)
        image_path = (get_backend_root() / 'media' / 'scanned_images' / image_filename).absolute()
        img_cv = cv2.imread(str(image_path))
        
        htr_engine = HTREngine()
        extracted_data = htr_engine.extract_exam_info(img_cv)
        
        res = {
            "method": "local_paddleocr",
            "score": extracted_data.get("score") if extracted_data else None,
            "studentId": extracted_data.get("studentId") if extracted_data else None,
            "studentName": extracted_data.get("studentName") if extracted_data else None,
            "confidence": extracted_data.get("confidence") if extracted_data else 0.0
        }
        
        logger.info(f"✅ [LOCAL-HTR-SUCCESS] Extracted: {res}")

        if res:
            # 🎯 Thực hiện Fuzzy Matching để chuẩn hoá ID học sinh
            roster_data = json.loads(config.get('roster_json', '[]'))
            matched_id, matched_name, is_match, score = find_best_student_match(
                res.get('studentId'), res.get('studentName'), roster_data
            )
            
            if is_match:
                res['studentId'] = matched_id
                res['studentName'] = matched_name
                res['isFuzzyMatch'] = True
                res['fuzzyScore'] = score
            else:
                res['isFuzzyMatch'] = False
                res['fuzzyScore'] = score

            # Extract mapping config for direct column index access
            mapping = config.get('mapping', {})
            id_col_idx = mapping.get('idCol') if mapping else None
            score_col_idx = mapping.get('scoreCol') if mapping else None
            data_row_start = mapping.get('dataRowStart', 2) if mapping else 2

            # Ghi điểm trực tiếp vào Excel (dùng ID đã map chuẩn)
            excel_updated = False
            if res.get('studentId') and res.get('score'):
                excel_updated = update_excel_score(
                    config['excel_filename'], config.get('sheet_name', ''),
                    res['studentId'], res['score'],
                    id_col_idx=id_col_idx, score_col_idx=score_col_idx, data_row_start=data_row_start
                )

            send_ws_update(session_id, "ai_result", {
                "status": "success",
                "studentId": res.get('studentId'),
                "studentName": res.get('studentName'),
                "score": res.get('score'),
                "isFuzzyMatch": res.get('isFuzzyMatch'),
                "fuzzyScore": res.get('fuzzyScore'),
                "excelUpdated": excel_updated,
                "image": f"/media/scanned_images/{image_filename}"
            })
        else:
            logger.warning(f"⚠️ [BG-AI-TASK] Gemini returned None for {image_filename}")
            send_ws_update(session_id, "ai_status", {"status": "error", "msg": "⚠️ AI không trả về kết quả."})
    except Exception as e:
        logger.error(f"💥 [BG-AI-TASK] Error: {str(e)}", exc_info=True)
        send_ws_update(session_id, "ai_status", {"status": "error", "msg": f"💥 Lỗi: {str(e)}"})

@api_view(['POST'])
@permission_classes([AllowAny])
@csrf_exempt
def scan_upload(request, session_id):
    img_b64 = request.data.get('image_base64', '')
    if img_b64:
        add_to_scan_store(session_id, img_b64)
        raw_bytes = extract_base64_data(img_b64)
        if raw_bytes:
            filename = f"mobile_{session_id}_{datetime.datetime.now().strftime('%H%M%S_%f')}.jpg"
            save_path = get_backend_root() / 'media' / 'scanned_images' / filename
            save_path.parent.mkdir(parents=True, exist_ok=True)
            with open(save_path, 'wb') as f: f.write(raw_bytes)

            config = get_session_config()
            if config.get("excel_filename"):
                # Sử dụng ThreadPoolExecutor thay vì threading.Thread trực tiếp
                ai_executor.submit(background_ai_task, filename, config.copy(), session_id)
            else:
                # Gửi thông báo cho User là ảnh đã lưu nhưng chưa có Excel để xử lý
                send_ws_update(session_id, "ai_status", {"status": "pending", "msg": "⏳ Ảnh đã lưu, vui lòng upload file Excel để bắt đầu chấm điểm."})
        return Response({"ok": True})
    return Response({"error": "Empty data"}, status=400)

@api_view(['GET'])
@permission_classes([AllowAny])
def scan_poll(request, session_id):
    imgs = pop_scan_store(session_id)
    return Response({"images": imgs, "count": len(imgs)})

@api_view(['GET'])
@permission_classes([AllowAny])
def server_info(request):
    return Response({"hostname": socket.gethostname(), "ai": "Gemini 1.5 Flash API + Pandas Core + OpenPyXL"})

@api_view(['GET'])
@permission_classes([AllowAny])
def get_sheet_results(request): return Response({"results": []})
@api_view(['POST'])
@permission_classes([AllowAny])
def delete_image(request): return Response({"success": True})
from django.http import FileResponse

@api_view(['GET'])
@permission_classes([AllowAny])
def download_updated_excel(request):
    filename = secure_filename(request.query_params.get('filename'))
    if not filename: return Response({"error": "No filename"}, status=400)
    
    with excel_lock:
        try:
            file_path = get_backend_root() / 'media' / 'rosters' / filename
            if not file_path.exists(): return Response({"error": "File not found"}, status=404)

            response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            logger.error(f"❌ [DOWNLOAD-EXCEL] Error: {str(e)}")
            return Response({"error": str(e)}, status=500)
